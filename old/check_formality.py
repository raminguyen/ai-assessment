import os
import json
import re
import glob
import csv
import pandas as pd

def get_informal_data(text):
    if not text:
        return False, "", ""
        
    text_lower = text.lower()
    
    # TRUNCATE BEFORE REFERENCES
    ref_headers = [
        r'(?m)^#*\s*references',
        r'(?m)^#*\s*bibliography',
        r'(?m)^#*\s*works cited',
        r'(?m)^#*\s*peer-reviewed references(\s*\(cited\))?',
        r'(?m)^#*\s*citations',
    ]
    
    clean_text = text
    clean_text_lower = text_lower
    for pattern in ref_headers:
        match = re.search(pattern, text_lower)
        if match:
            clean_text = text[:match.start()]
            clean_text_lower = text_lower[:match.start()]
            break
    
    # Pronouns
    # First person: Exclude 'i' when used as roman numeral (i), initial (i.), or range (i-)
    first_person = [r'(?<!\()\bi\b(?![.\-\u2013)\)])', r'\bmyself\b', r'\bmy\b', r'\bmine\b', r'\bus\b']
    # Second person
    second_person = [r'\byou\b', r'\byour\b', r'\byours\b']
    
    # Step keywords / Enumerations
    steps = [r'\bstep 1\b', r'\bstep 2\b']

    # Remove placeholders and quoted assignment text before checking
    exclude_patterns = [
        r'\[your name\]',
        r'fine,?\s*fine,?\s*i[\'\u2019]ll feed you now',
        r'i like this greeting,?\s*so i keep doing it',
        r'when would you \*?not\*? need social learning to explain the scene\??\*?',
        r'i can[\'\u2019]t literally .?think step-by-step.? in the sense of revealing private internal reasoning,?\s*but i \*?can\*? present a clear,?\s*structured,?\s*evidence-based analysis with the logic made explicit\.?',
        r'the cats are not reasoning .?i should go to the kitchen to get food.?;?\s*they are responding automatically to a conditioned stimulus\.?',
        r'who is .?them,.?\s*and what claim or misconception am i setting them straight about.*?(?:etc\.\)?\s*\??|risk-free,.?\s*etc\.?\)?\s*\??)',
        r'i can[\'\u2019]t write the full essay for you to submit as-is\.?',
        r'i \*?can\*? help you .?set them straight.?.*?you can customize into your own[^.]*\.?',
        r'below is a ready-to-adapt essay you can revise into your own words[^.]*\.?',
        r'pavlovian conditioning:?\s*it[\'\u2019]s not what you think it is\.?',
        r'if your instructor requires \*?articles only\*?[^.]*\.?',
        r'tell me whether your class allows books\.?',
        r'if you want,?\s*i can tailor it\.?',
        r'reply with:?\s*1\)?\s*whether \*?\*?books are allowed\*?\*?.*?you can rewrite and submit\.?',
        r'i[\'\u2019]ll adapt the essay into a final[^.]*\.?',
        r'note about your .?peer.?reviewed articles.? requirement.*?(?:use:|options)',
        r'i can suggest additional article-only options',
        r'here is the essay drafted based on your discussion,?\s*designed to clarify these concepts for your study group\.?',
        r'social learning in dog training:?\s*the effectiveness of the do as i do method compared to shaping/?clicker training\.?',
        r'who is .?them,.?\s*and what claim\(?s?\)?\s*do you want to correct',
        r'what format do you need \(?apa/?mla/?chicago\)?\s*and should the essay focus strictly on[^?]*\??',
        r'in the meantime,?\s*here is a[^.]*essay[^.]*that .?sets them straight.?[^.]*\.?',
        r'i[\'\u2019]ll format citations in \*?\*?apa[\-\u2011]?style\*?\*?\s*and use peer[\-\u2011\u2010]?reviewed sources\.?',
        r'i can also format citations in apa or mla\.?',
        r'i can \(?a\)?\s*swap in[^.]*format citations in apa[^.]*\.?',
        r'to do this correctly i need one detail:?',
        r'that boston real estate is basically a guaranteed win because demand is strong,?\s*so investors can ignore interest rates,?\s*supply constraints,?\s*and regulation',
    ]
    for pat in exclude_patterns:
        clean_text_lower = re.sub(pat, '', clean_text_lower)
        clean_text = re.sub(pat, '', clean_text, flags=re.IGNORECASE)

    combined_pattern = '|'.join(first_person + second_person + steps)
    
    matches = list(re.finditer(combined_pattern, clean_text, flags=re.IGNORECASE))
    
    if not matches:
        return False, "", ""
        
    evidences = []
    found_keywords = []
    
    for match in matches:
        start_idx = match.start()
        end_idx = match.end()
        
        # Get the word in original casing
        matched_word = match.group(0)
        
        # If it's 'US' (all caps), it's likely the acronym for United States or Unconditioned Stimulus
        if matched_word == "US":
            continue
            
        # Approximate sentence extraction
        sent_start = max(0, clean_text.rfind('.', 0, start_idx) + 1)
        sent_end = clean_text.find('.', end_idx)
        if sent_end == -1:
            sent_end = len(clean_text)
        else:
            sent_end += 1
            
        sentence = clean_text[sent_start:sent_end].strip().replace('\n', ' ')
        
        if sentence and sentence not in evidences:
            evidences.append(sentence)
        
        # Capture the actual matched word in lowercase for reporting
        matched_word_lower = matched_word.lower()
        if matched_word_lower not in found_keywords:
            found_keywords.append(matched_word_lower)
            
    if not evidences:
        return False, "", ""

    return True, ", ".join(found_keywords), " | ".join(evidences)

def main():
    assignments = [1, 2, 3]
    models = ['chatgpt', 'claude', 'gemini', 'grok']
    prompts = range(0, 12)

    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    data_dir = os.path.join(base_dir, 'data', 'critical_thinking')
    if not os.path.exists(data_dir):
        data_dir = os.path.join('data', 'critical_thinking')

    results = {} # (a, m, p, essay_type) -> (is_informal, keywords, evidence)
    word_counts = {} # (a, m, p) -> word count
    files = glob.glob(os.path.join(data_dir, 'a*_gen_*.json')) + glob.glob(os.path.join(data_dir, 'a*_tune_*.json'))

    for filepath in files:
        filename = os.path.basename(filepath)
        match = re.match(r'a(\d+)_(gen|tune)_([a-z]+)_p(\d+)\.json', filename)

        if match:
            a_num = int(match.group(1))
            essay_type = 'generate' if match.group(2) == 'gen' else 'tune'
            m_name = match.group(3)
            p_num = int(match.group(4))

            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    text_to_check = data.get('result', '')
                    results[(a_num, m_name, p_num, essay_type)] = get_informal_data(text_to_check)
                    if essay_type == 'generate':
                        word_counts[(a_num, m_name, p_num)] = len(text_to_check.split())
            except Exception:
                pass

    # Write combined results to CSV with evidence
    evidence_csv = 'formality_results_with_evidence.csv'
    with open(evidence_csv, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['Assignment', 'EssayType', 'Model', 'Prompt', 'Is Informal', 'Matched Keywords', 'Exact Sentence'])

        for a in assignments:
            for m in models:
                for p in prompts:
                    for essay_type in ['generate', 'tune']:
                        is_inf, keywords, evidence = results.get((a, m, p, essay_type), (None, "", ""))
                        if is_inf is None:
                            continue
                        status = 'x' if is_inf else ' '
                        writer.writerow([f'Assignment {a}', essay_type, m, f'p{p}', status, keywords, evidence])
    
    print(f"Detailed results with evidence written to {evidence_csv}")

    # Write summary table to CSV
    summary_csv = 'formality_check_results.csv'
    with open(summary_csv, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        
        # Header row
        header = ['Assignment', 'Model', 'EssayType'] + [f'p{p}' for p in prompts]
        writer.writerow(header)

        # Technique name row (second row)
        technique_names = {
            0: 'Write',
            1: 'Persona + Write',
            2: 'Persona + Write Formal',
            3: 'CoT + Write',
            4: 'Generate',
            5: 'CoT + Generate',
            6: 'Persona + Generate',
            7: 'Iterative + Persona + Generate',
            8: 'Iterative + Persona + Write',
            9: 'Iterative + Generate',
            10: 'Iterative + CoT + Generate',
            11: 'Iterative + Write'
        }
        technique_row = ['Technique', '', '']
        for p in prompts:
            technique_row.append(technique_names.get(p, ''))
        writer.writerow(technique_row)

        # Initialize counts
        counts = {p: 0 for p in prompts}
        assignment_counts = {a: {p: 0 for p in prompts} for a in assignments}

        for a in assignments:
            for m in models:

                gen_row = [f'Assignment {a}', m, 'generate']
                tune_row = [f'Assignment {a}', m, 'tune']

                for p in prompts:
                    is_inf_gen, _, _ = results.get((a, m, p, 'generate'), (False, "", ""))
                    is_inf_tune, _, _ = results.get((a, m, p, 'tune'), (False, "", ""))

                    gen_row.append('x' if is_inf_gen else ' ')
                    tune_row.append('x' if is_inf_tune else ' ')

                    if is_inf_gen:
                        counts[p] += 1
                        assignment_counts[a][p] += 1

                writer.writerow(gen_row)
                writer.writerow(tune_row)

            # Add per-assignment total row
            a_total = ['', f'Total A{a}', '']
            for p in prompts:
                a_total.append(str(assignment_counts[a][p]))
            writer.writerow(a_total)

        # Add overall total row
        total_row = ['Total Informal Essay', '', '']
        for p in prompts:
            total_row.append(str(counts[p]))
        writer.writerow(total_row)

        # Add avg total score rows from scores_all.csv
        writer.writerow([])

        avg_gen_row = ['Avg Total Score (Generate)', '', '']
        avg_tune_row = ['Avg Total Score (Tune)', '', '']

        scores_all_path = 'scores_all.csv'

        if os.path.exists(scores_all_path):

            df_all = pd.read_csv(scores_all_path)
            df_all['Total'] = pd.to_numeric(df_all['Total'], errors='coerce')

            for p in prompts:
                p_df = df_all[df_all['Prompt'] == 'p' + str(p)]

                gen = p_df[p_df['EssayType'] == 'generate']['Total']
                tune = p_df[p_df['EssayType'] == 'tune']['Total']

                avg_gen_row.append(str(round(gen.mean(), 2)) if len(gen) > 0 else '')
                avg_tune_row.append(str(round(tune.mean(), 2)) if len(tune) > 0 else '')

        else:

            for p in prompts:
                avg_gen_row.append('')
                avg_tune_row.append('')

        writer.writerow(avg_gen_row)
        writer.writerow(avg_tune_row)

    print(f"Summary results written to {summary_csv}")

    # Write word count CSV
    
    wc_csv = 'word_count_results.csv'
    with open(wc_csv, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        header = ['Assignment', 'Model'] + [f'p{p}' for p in prompts]
        writer.writerow(header)

        technique_row = ['Technique', '']
        for p in prompts:
            technique_row.append(technique_names.get(p, ''))
        writer.writerow(technique_row)

        for a in assignments:
            for m in models:
                row = [f'Assignment {a}', m]
                for p in prompts:
                    wc = word_counts.get((a, m, p), '')
                    row.append(wc)
                writer.writerow(row)

    print(f"Word count results written to {wc_csv}")

    # Write Excel file with both sheets

    excel_file = 'formality_results.xlsx'

    df_evidence = pd.read_csv(evidence_csv)

    df_summary = pd.read_csv(summary_csv)

    df_wc = pd.read_csv(wc_csv)

    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df_evidence.to_excel(writer, sheet_name='Evidence', index=False)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        df_wc.to_excel(writer, sheet_name='Word Count', index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    wb = load_workbook(excel_file)
    ws = wb['Summary']
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    for row in ws.iter_rows():
        if row[0].value in ('Avg Total Score (Generate)', 'Avg Total Score (Tune)'):
            score_cells = []
            for c in row:
                if c.value is not None and c.value != '' and c.column > 2:
                    score_cells.append((c, float(c.value)))
            if score_cells:
                max_val = max(v for _, v in score_cells)
                for c, v in score_cells:
                    if v == max_val:
                        c.fill = red_fill

    wb.save(excel_file)
    print(f"Excel results written to {excel_file}")

if __name__ == "__main__":
    main()
